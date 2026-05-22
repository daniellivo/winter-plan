#!/usr/bin/env python3
"""Parse the Neonatos xlsx into a flat list of {date, slot, count} entries."""

import json
import sys
from datetime import datetime
from openpyxl import load_workbook

XLSX = '/Users/dani/Downloads/Verano_Livo_2026 NEONATOS.xlsx'
OUT  = '/Users/dani/Desktop/winter-plan/scripts/neonatos-shifts.json'

SLOT_BY_KEYWORD = {
    'mañanas': 'TM',
    'mananas': 'TM',
    'tardes':  'TT',
    'noches':  'TN',
}

MONTH_HEADERS = {'JUNIO', 'JULIO', 'AGOSTO', 'SEPTIEMBRE'}  # Stop after September

wb = load_workbook(XLSX)
ws = wb['UCI NEONATAL']

rows = list(ws.iter_rows(values_only=True))
out = []
seen_first_header = False
done = False

i = 0
while i < len(rows) and not done:
    row = rows[i]
    first = (row[0] or '')
    first_str = str(first).strip() if first is not None else ''

    if first_str.upper() in MONTH_HEADERS:
        seen_first_header = True
        i += 1
        continue
    if first_str.upper() == 'OCTUBRE':
        done = True
        break

    if not seen_first_header:
        i += 1
        continue

    # Date row: col A empty, col B empty, cols C..I have datetimes.
    if first_str == '' and (row[1] is None or str(row[1]).strip() == ''):
        dates = []
        for j in range(2, 9):
            cell = row[j] if j < len(row) else None
            if isinstance(cell, datetime):
                # Filter to 2026 only — sheet has a 2024 example block and a 2025 stray block.
                if cell.year == 2026 and 6 <= cell.month <= 9:
                    dates.append(cell.strftime('%Y-%m-%d'))
                else:
                    dates.append(None)
            else:
                dates.append(None)
        if any(d is not None for d in dates):
            # Next 3 rows = TM, TT, TN (by col B label).
            for k in range(1, 4):
                if i + k >= len(rows):
                    break
                slot_row = rows[i + k]
                label = (slot_row[1] or '')
                label_str = str(label).strip().lower() if label is not None else ''
                slot = SLOT_BY_KEYWORD.get(label_str)
                if slot is None:
                    continue
                for idx, d in enumerate(dates):
                    if d is None:
                        continue
                    val = slot_row[2 + idx] if 2 + idx < len(slot_row) else None
                    if val is None:
                        continue
                    try:
                        n = int(val)
                    except (TypeError, ValueError):
                        continue
                    if n <= 0:
                        continue
                    out.append({'date': d, 'slot': slot, 'count': n})
            i += 4
            continue
    i += 1

# Dedup keep max
dedup = {}
for e in out:
    key = (e['date'], e['slot'])
    if key not in dedup or e['count'] > dedup[key]['count']:
        dedup[key] = e
final = sorted(dedup.values(), key=lambda x: (x['date'], {'TM': 0, 'TT': 1, 'TN': 2}[x['slot']]))

print(f'Total entries: {len(final)}', file=sys.stderr)
total_shifts = sum(e['count'] for e in final)
print(f'Total physical shifts: {total_shifts}', file=sys.stderr)

with open(OUT, 'w') as f:
    json.dump(final, f, indent=2, ensure_ascii=False)
print(f'Wrote {OUT}', file=sys.stderr)
